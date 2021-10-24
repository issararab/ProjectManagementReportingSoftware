#-- ========================================================================
#-- Author      : <Issar Arab>
#-- Created     : <date: 21.09.2020>
#-- Description : <Script to pull customized reports for the hours worked by
#--                all your team members in projects encoded in Toggl. 
#--                The output table summarizes the amount of time spent 
#--                in each task and helps the project managers/team leads 
#--                mange and keep track their teams performance.>
#-- ========================================================================

import pandas as pd
import numpy as np
from toggl.TogglPy import Toggl
import datetime
# import multiprocessing
from joblib import Parallel, delayed
import threading
import itertools
import time
import math
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

df_global = pd.DataFrame()

def animate():
    while True:
        for c in itertools.cycle(['|', '/', '-', '\\']):
            sys.stdout.write('\rPulling: ' + c)
            sys.stdout.flush()
            time.sleep(0.1)


with open('Toggl_Config.json') as config_file:
    config = json.load(config_file)


def RepresentsInt(s):
    try: 
        int(s)
        return True
    except ValueError:
        return False
        
        
def gen_custom_report(content):
    flatten_list = []
    for o in content:
        for df in o:
            flatten_list.append(df)
    report_df = pd.concat(flatten_list)
    #Copy Extracted information to Global DataFrame
    global df_global
    df_global = report_df
    # Format datatime
    dates = report_df[['start']].copy()
    dates = dates['start'].apply(lambda x: str(x)[:10])
    fiscal_month = ['01.{:02d}.{}'.format(int(m), int(y)) for y, m, d in map(lambda x: x.split('-'), dates)]
  
    duration_in_hours = np.round(report_df['dur'].values / (1000*3600), 3)
    # Generate final dataframe
    df_to_export = report_df[['user']].copy()
    df_to_export.insert(1, 'fiscal_month', fiscal_month)
    df_to_export.insert(2, 'project', report_df['project'])
    df_to_export.insert(3, 'task', report_df['task'])
    df_to_export.fillna('-', inplace=True)
    df_to_export.insert(4, 'tags', report_df['tags'].apply(lambda x: ', '.join(x)))
    df_to_export.insert(5, 'Duration in Hours', duration_in_hours)
    df_to_export.insert(6, 'Is Billable', report_df['is_billable'])
    # Time rounding to the closest quarter
    df_to_export.insert(7, 'time_tracked_in_hours', df_to_export['Duration in Hours'].apply(lambda x: round(x*4)/4))
    df_to_export = df_to_export.drop(columns=['Duration in Hours'])
    
    # Generate fiscal year report
    pivot_list = ['user', 'fiscal_month', 'project', 'task', 'tags', 'Is Billable']
    df_to_export = df_to_export.groupby(pivot_list)['time_tracked_in_hours'].sum().reset_index()
    
    # Sort by datetime decending
    df_to_export.sort_values(['user', 'fiscal_month'], ascending=[True, True], inplace=True)

    # Create kva_num
    df_to_export.project.apply(str)
    df_to_export.time_tracked_in_hours.apply(str)
    df_to_export.project.fillna(value='', inplace=True)
    df_to_export.insert(2, 'ProjectName', df_to_export['project'].apply(lambda x: str(x.split('-', 2)[-1].strip())))
    df_to_export.insert(2, 'TeamID', df_to_export['project'].apply(
        lambda x: str(x.split('-', 2)[1].strip())
        if len(x.split('-', 2)) == 3 and RepresentsInt(x.split('-', 2)[1].strip())
        else math.nan))
    df_to_export.insert(2, 'KVA-Nr', df_to_export['project'].apply(
        lambda x: str(x.split('-', 1)[0].strip())
        if RepresentsInt(x.split('-', 1)[0].strip())
        else "3"))  # math.nan
    df_to_export.loc[df_to_export['task'] == 'CK Dynamics Project', 'KVA-Nr'] = "1"
    df_to_export.loc[df_to_export['task'] == 'Partner Scoring', 'KVA-Nr'] = "2"
    # df_to_export.rename(columns={'project':'ProjectName'}, inplace=True)
    df_to_export = df_to_export.drop(columns=['project'])
    return df_to_export


def export_to_csv(content):
    content.to_csv('FY'+str(config['fiscal_year'])[:-2]+'_toggl_monthly'+'.csv', index=False)  # , sep=';')


def export_to_excel(content):
    content.to_excel('FY'+str(config['fiscal_year'])[:-2]+'_toggl_monthly'+'.xlsx', index=False)  # , sep=';')

def generate_KVA_excels():
    # get Data from Global DataFrame for local Iteration
    df_kva = df_global
    # adding KVA-Nr as new Column to dataframe
    df_kva.project.apply(str)
    df_kva.project.fillna(value='', inplace=True)
    df_kva["KVA-Nr"]= df_kva['project'].apply(
        lambda x: str(x.split('-', 1)[0].strip())
        if RepresentsInt(x.split('-', 1)[0].strip())
        else "3")
    df_kva["KVA-Nr"]= pd.to_numeric(df_kva["KVA-Nr"])    
    df_kva["start"]= pd.to_datetime(df_kva["start"])
    df_kva["end"]= pd.to_datetime(df_kva["end"])
    df_kva["Start Date"] = df_kva["start"].dt.date
    df_kva["Start Time"] = df_kva["start"].dt.time
    df_kva["Stop Date"] = df_kva["end"].dt.date
    df_kva["Stop Time"] = df_kva["end"].dt.time
    duration_in_hours = df_kva['dur'].values // (1000*3600)
    df_kva["tags"]= df_kva['tags'].apply(lambda x: ', '.join(x))
    df_kva["Time(decimals)"]= df_kva["dur"].apply(lambda x: round(np.round(x/(1000*3600), 3)*4)/4)
    
    # get Month value from Config File. If Set to 'Last Month', get values for Last Month otherwise get the values from the String. 
    # Month Format 'MM-YYYY'
    get_month = config['Fiscal_month']
    if(get_month == "Last Month"):
        today = datetime.datetime.today()
        first = today.replace(day=1)
        get_month = first - datetime.timedelta(days=1) 
    else:   
        get_month =datetime.datetime.strptime(get_month, "%m/%Y")

    df_kva = df_kva[df_kva["start"].dt.month == get_month.month]
    KVA_Nums = config['KVA']

    for x in KVA_Nums:
        #generate excel specific KVA numbers
        df = df_kva[df_kva['KVA-Nr']==x]
        df = df.fillna("")
        
        #Get Project Name
        projectname = ''.join(df["project"].unique())
        projectname =projectname.split('-', 2)[-1].strip()

        #Build Separte Dataframe to select columns
        df_to_export = df[['project','task','tags','Start Date', "Time(decimals)", 'dur']].copy()

        # Generate fiscal year report by Grouping on Time Columns
        pivot_list = ['project','task','tags','Start Date']
        df_to_export = df_to_export.groupby(pivot_list)[['dur', 'Time(decimals)']].sum().reset_index()
        df_to_export.rename(columns={'Start Date': 'Date'}, inplace=True)
        df_to_export["Time(h)"] = df_to_export['dur'].apply( lambda x: millisecondsToHHMMSS(x))
        df_to_export = df_to_export.drop(columns=['dur'])
        df_to_export.columns = map(str.upper, df_to_export.columns)   

        #Excel Worksheet Formation
        if not df_to_export.empty:
            wb = Workbook()
            ws = wb.active
            fillBlue = PatternFill(start_color='0099CCFF',
                    end_color='0099CCFF',
                    fill_type='solid')
            ws['B2']= "Month"
            ws['B2'].fill = fillBlue
            ws['B3']= "KVA-Nr"
            ws['B3'].fill = fillBlue
            ws['C2']=get_month.strftime("%B %Y")
            ws['C2'].fill = fillBlue
            ws['C3']= str(x)
            ws['C3'].fill = fillBlue
            rows = len(df_to_export.index) + 5
            for r in dataframe_to_rows(df_to_export, index=False, header=True):
                ws.append(r)
            
            #Table Styling Code Below
            ws.insert_rows(4)
            tab = Table(displayName="Table1", ref="A5:F"+ str(rows))
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab) 
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
            ws.column_dimensions = dim_holder                  
            wb.save(str(x)+ "_"+ str(get_month.date())[0:7]+"_POE_"+projectname+".xlsx")
        


def millisecondsToHHMMSS(ms):
    seconds = math.floor((ms / 1000) % 60)
    minutes = math.floor((ms / 1000 / 60) % 60)
    hours = math.floor((ms  / 1000 / 3600 ) % 24)

    hours = "0" + str(hours) if hours < 10    else str(hours)
    minutes = "0" + str(minutes) if minutes < 10 else str(minutes)
    seconds = "0" + str(seconds) if seconds < 10 else str(seconds)

    return hours + ":" + minutes + ":" + seconds
    

def pull_data(limits, token, parameters):
    start = limits[0]
    end = limits[1]
    toggl = Toggl()
    toggl.setAPIKey(token)
    list_of_dfs = []
    while start <= end:
        parameters['page'] = start
        detailed_report = toggl.request(detailed_report_url, parameters)
        start += 1
        list_of_dfs.append(pd.DataFrame(detailed_report['data']))
    # output_list.append(list_of_dfs)
    return list_of_dfs


def authenticate_and_get_workspace_content(token, user_email):
    print("Authenticated.")
    time.sleep(0.5)
    toggl = Toggl()
    toggl.setAPIKey(token)
    curr_date = datetime.datetime.now()
    if config['current_fiscal_year']:
        curr_year = curr_date.year
        curr_month = curr_date.month
        curr_day = curr_date.day
    else:
        curr_year = config['fiscal_year']
        curr_month = 6
        curr_day = 30
    print("Connected.")
    time.sleep(0.5)
    creakom_obj = toggl.getWorkspace(name='Creakom')
    parameters = {
        'user_agent': user_email,
        'workspace_id': creakom_obj['id'],
        'since': str((curr_year - 1) if int(curr_month) < 7 else curr_year) + '-' + str(7) + '-' + str(1),
        'until': str(curr_year) + '-' + str(curr_month) + '-' + str(curr_day),
        "order_field": "user",
        "page": 1
    }
    t = threading.Thread(target=animate)
    t.daemon = True
    t.start()
    # Initial request to know the no. of pages to load
    detailed_report = toggl.request(detailed_report_url, parameters)
    # No. of workers calculations
    total_records = detailed_report['total_count']
    per_page = detailed_report['per_page']
    number_of_pages = math.ceil(total_records/per_page)
    cores = 1  # math.ceil(multiprocessing.cpu_count()/2)
    # at least each core should handle 10 requests
    cores = min(cores, max(math.floor(number_of_pages/10), 1))
    min_val = 1
    max_val = math.ceil(number_of_pages/cores)
    pages_per_thread = max_val
    pages_dist = []
    for i in range(cores):
        pages_dist.append((min_val, max_val))
        min_val = max_val + 1
        max_val = min(max_val+pages_per_thread, number_of_pages)
    # output_list = parallel_processing(pages_dist[0], token, parameters)
    output_list = list(Parallel(n_jobs=cores)(delayed(pull_data)(val, token, parameters) for val in pages_dist))
    
    return output_list


if __name__ == "__main__":
    api_token = "Enter API Token: "
    token = config['token']
    while len(token) == 0:
        print(api_token)
        token = input()
    email_msg = "Enter Your Email, preferabe the one used for Toggl: "
    user_email = config['user_email']
    while len(user_email) == 0:
        print(email_msg)
        user_email = input()
    tic = datetime.datetime.now()
    detailed_report_url = 'https://toggl.com/reports/api/v2/details'
    try:
        workspace_content = authenticate_and_get_workspace_content(token, user_email)
    except Exception as e:
        print('Error, please check if your token is valid!')
        sys.exit(0)
    export_to_excel(gen_custom_report(workspace_content))
    generate_KVA_excels()
    toc = datetime.datetime.now()
    done = True
    print("\rDone      !\nIt took {} secs.".format((toc-tic).total_seconds()))
sys.exit(0)
