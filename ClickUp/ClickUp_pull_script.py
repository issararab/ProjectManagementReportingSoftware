#-- ========================================================================
#-- Author      : <Issar Arab>
#-- Created     : <date: 20.01.2021>
#-- Description : <Script to pull all closed tasks and sub-tasks
#--                in a given month for a particular fiscal year 
#--                from ClickUp. Script allows dynamic selection of 
#--                columns to output. You just need to provide the token
#--                in the config file.>
#-- ========================================================================

from pyclickup import ClickUp
import json
import datetime
import sys
import pandas as pd
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

def animate():
    while True:
        for c in itertools.cycle(['|', '/', '-', '\\']):
            sys.stdout.write('\rPulling: ' + c)
            sys.stdout.flush()
            time.sleep(0.1)

def get_Taskname(listOfTasks, parentID, i):
    for t in listOfTasks[i:]:
        if parentID == t.id:
            return t.name
    return None
    
def return_entry_object(team,sp,prot,lst,task):
    tags = 'None' if not task.tags else '; '.join([tag.name for tag in task.tags])
    priority = 'None' if task.priority is None else task.priority['priority']
    users = 'None' if task.assignees is None else '; '.join([str(assignee.username) for assignee in task.assignees])
    entry = {'Team': team.name,
             'Space': sp.name,
             'Folder': prot.name,
             'List': lst.name,
             'Task ID': task.id,
             'Task': task.name if task.parent is None else get_Taskname(lstTasks,task.parent,counter),
             'Sub-Task': "" if task.parent is None else task.name,
             'Tags': tags,
             'Priority': priority,
             'Status': task.status.status,
             'Assignees': users, 
             'Due date': str(task.due_date.day)+'/'+str(task.due_date.month)+'/'+str(task.due_date.year) if task.due_date != None else "",
             'Status modified date': str(task.date_updated.day)+'/'+str(task.date_updated.month)+'/'+str(task.date_updated.year)}
    return entry
    
def generate_kva(df):
    month = config['month'] 
    year = config['year']  
    df_kva = df
    if (month > 6):
        FiscalYear= "FY"+ str(year+1)[-2:0]
    else:
        FiscalYear= "FY"+ str(year)[-2:0]

    for folder in config['folders_kva']:
        folder_name = folder['folder name']
        kva_nr = folder['kva_nr']          
 
        #generate excel specific KVA numbers
        df = df_kva[df_kva['Folder']==folder_name]
        df = df.fillna("")                        

        #Build Separte Dataframe to select columns
        df_to_export = df[['List','Task','Sub-Task','Tags', "Priority", 'Status modified date']].copy() 

        #Excel Worksheet Formation
        if not df_to_export.empty:
            wb = Workbook()
            ws = wb.active
            fillBlue = PatternFill(start_color='0099CCFF',
                    end_color='0099CCFF',
                    fill_type='solid')
            ws['B2']= "Month"
            ws['B2'].fill = fillBlue
            ws['B3']= "KVA-Nr"
            ws['B3'].fill = fillBlue
            ws['C2']=str(month)+"/" +str(year)
            ws['C2'].fill = fillBlue
            ws['C3']= str(kva_nr)
            ws['C3'].fill = fillBlue
            rows = len(df_to_export.index) + 5
            for r in dataframe_to_rows(df_to_export, index=False, header=True):
                ws.append(r)
            
            #Table Styling Code Below
            ws.insert_rows(4)
            tab = Table(displayName="Table1", ref="A5:F"+ str(rows))
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab) 
            dim_holder = DimensionHolder(worksheet=ws)
            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
            ws.column_dimensions = dim_holder                  
            wb.save(str(kva_nr)+"_POE_"+ str(folder_name).replace(' ','')+"_"+FiscalYear+'_' +str(month)+".xlsx")
            print("###### Generated Clickup KVA"+str(kva_nr)+" Excel #######")


if __name__ == "__main__":
    #Data pull config
    with open('ClickUp_Config.json') as config_file:
        config = json.load(config_file)

    if len(config['token']) == 0:
        print('Check ClickUp_Config.json! \nMake sure you have a none empty token in your Json config file.')
        sys.exit(0)
    
    #Instantiate API conection
    try:
        clickup = ClickUp(config['token'])
    except Exception as e:
        print('Error, please check if your token is valid!')
        sys.exit(0)
    
    entries = []
    tic = datetime.datetime.now()  
    #Pull data
    for team in clickup.teams:
        for sp in team.spaces:
            for prot in sp.projects:#Folder
                if config['folder'] != 'All' and prot.name != config['folder']:
                    continue
                for lst in prot.lists:
                    if config['list'] != 'All' and lst.name != config['list']:
                        continue
                    lstTasks = lst.get_all_tasks(include_closed=config['closed_tasks'],subtasks=True)
                    for counter, task in enumerate(lstTasks):
                        if task.parent is not None and get_Taskname(lstTasks,task.parent,counter) is None:
                            continue
                        if config['closed_tasks']:
                            if task.status.status == 'Closed' and task.date_updated.month == config['month'] and task.date_updated.year == config['year']:
                                entry = return_entry_object(team,sp,prot,lst,task)
                                # print(entry)
                                entries.append(entry)
                        else:
                            entry = return_entry_object(team,sp,prot,lst,task)
                            # print(entry)
                            entries.append(entry)
                    sleep(1)
                                
                            
    #Write objects to data frame
    output_table = pd.DataFrame(entries)
    output_table.sort_values(['Task', 'Sub-Task'], inplace=True)
    
    

    #Select the fiels to output
    for field in config['fields']:
        if not field[next(iter(field))]:
            output_table.drop(columns=[next(iter(field))], inplace=True)
    
    #Write the results in excel
    output_table.to_excel(config['output_file'], index= False) 
    print("###### Generated Clickup Output Excel #######")
    #Generate KVA excels for ClickUps
    generate_kva(output_table)

    toc = datetime.datetime.now()
    print("\rDone      !\nIt took {} secs.".format((toc-tic).total_seconds()))

sys.exit(0)

